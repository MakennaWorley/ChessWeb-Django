from openpyxl import load_workbook
from openpyxl.styles.fonts import Font

from .models import Player, Game
import os
from django.conf import settings
from datetime import datetime


def write_ratings():
    file_path = os.path.join(settings.BASE_DIR, 'files', 'RatingsTemplate.xlsx')
    workbook = load_workbook(file_path)
    sheet = workbook.active

    students = Player.objects.filter(active_member=True, is_volunteer=False, is_active=True).order_by('-rating',
                                                                                                      '-grade',
                                                                                                      'last_name',
                                                                                                      'first_name')

    start_row = 2

    for index, student in enumerate(students, start=start_row):
        sheet[f'A{index}'] = student.name()  # Name
        sheet[f'B{index}'] = student.grade  # Grade
        sheet[f'C{index}'] = student.rating  # Rating
        sheet[f'D{index}'] = student.lesson_class.name if student.lesson_class else ''

    date = datetime.now().strftime('%m-%d-%Y')

    new_file_path = os.path.join(settings.BASE_DIR, 'files', 'ratings', f'Ratings_{date}.xlsx')
    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)
    workbook.save(new_file_path)

    return new_file_path


def write_pairings(submitted_date):
    file_path = os.path.join(settings.BASE_DIR, 'files', 'PairingTemplate.xlsx')
    workbook = load_workbook(file_path)
    sheet = workbook.active

    bold_font = Font(bold=True)

    games = Game.objects.filter(date_of_match=submitted_date, is_active=True)

    for game in games:
        white_player = game.white.name() if game.white else ""
        black_player = game.black.name() if game.black else ""
        board = game.get_board()

        matching_row = None
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if row[0] == board:
                matching_row = row_index
                break

        if matching_row:
            white_cell = sheet[f'B{matching_row}']
            white_cell.value = white_player

            if game.white and game.white.is_volunteer:
                white_cell.font = bold_font

            black_cell = sheet[f'D{matching_row}']
            black_cell.value = black_player

            if game.black and game.black.is_volunteer:
                black_cell.font = bold_font

        else:
            print(f"No matching board found for game: {game}")

    new_file_path = os.path.join(settings.BASE_DIR, 'files', 'pairings', f'Pairings_{submitted_date}.xlsx')
    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)
    workbook.save(new_file_path)

    return new_file_path


def export_player_data():
    file_path = os.path.join(settings.BASE_DIR, 'files', 'ExportTemplate.xlsx')
    workbook = load_workbook(file_path)
    sheet = workbook.active

    students = Player.objects.filter(active_member=True, is_volunteer=False, is_active=True).order_by('-rating',
                                                                                                      '-grade',
                                                                                                      'last_name',
                                                                                                      'first_name')

    start_row = 2

    for index, student in enumerate(students, start=start_row):
        sheet[f'A{index}'] = student.name()
        sheet[f'B{index}'] = student.rating
        sheet[f'C{index}'] = student.beginning_rating
        sheet[f'D{index}'] = student.improved_rating()
        sheet[f'E{index}'] = student.grade
        sheet[f'F{index}'] = student.lesson_class.name if student.lesson_class else ''
        sheet[f'G{index}'] = student.parent_or_guardian
        sheet[f'H{index}'] = student.email
        sheet[f'I{index}'] = student.phone

    date = datetime.now().strftime('%m-%d-%Y')

    new_file_path = os.path.join(settings.BASE_DIR, 'files', 'ratings', f'Export_{date}.xlsx')
    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)
    workbook.save(new_file_path)

    return new_file_path